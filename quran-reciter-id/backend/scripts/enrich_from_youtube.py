"""
Enrich reciter database from YouTube.

For each reciter in the Excel that is NOT yet in the fingerprint DB:
  1. Search YouTube for `<name> قرآن`
  2. Download top-N results as audio (up to 60s each)
  3. Compute ensemble embedding via the loaded AI engine
  4. Store as `Name#i` fingerprints (multi-fingerprint support)

Uses yt-dlp (pure Python, no cookies) — best-effort per reciter.
Skips reciters where all downloads fail.

Env:
  YT_PER_RECITER   default 3  (how many fingerprints per reciter)
  YT_CLIP_SECONDS  default 60
  QURAN_EXCEL      default backend/data/reciters.xlsx
"""
import json
import os
import re
import sys
import tempfile
import time
import unicodedata
from pathlib import Path

BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND))

DB_PATH = BACKEND / "data" / "embeddings" / "reciter_database.json"
META_PATH = BACKEND / "data" / "reciters_metadata.json"
EXCEL_PATH = Path(os.environ.get("QURAN_EXCEL", BACKEND / "data" / "reciters.xlsx"))
PER_RECITER = int(os.environ.get("YT_PER_RECITER", "3"))
CLIP_SECONDS = int(os.environ.get("YT_CLIP_SECONDS", "60"))
CLIP_OFFSET = 10  # skip intro


def _norm(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"[إأآا]", "ا", s)
    s = re.sub(r"[ىي]", "ي", s)
    s = re.sub(r"ة", "ه", s)
    s = re.sub(r"[ًٌٍَُِّْـ]", "", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _load_existing_bases() -> set:
    if not DB_PATH.exists():
        return set()
    with open(DB_PATH, encoding="utf-8") as f:
        db = json.load(f)
    bases = set()
    for r in db.get("reciters", []):
        n = r["reciter_name"].split("#")[0]
        bases.add(_norm(n))
    return bases


def _load_excel_names() -> list:
    if not EXCEL_PATH.exists():
        print(f"[skip] excel not found at {EXCEL_PATH}", flush=True)
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[skip] openpyxl not installed", flush=True)
        return []
    wb = load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active
    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and len(cell.strip()) > 2:
                names.append(cell.strip())
                break
    # dedupe
    seen, out = set(), []
    for n in names:
        k = _norm(n)
        if k and k not in seen:
            seen.add(k); out.append(n)
    return out


def _yt_search_and_download(name: str, per: int, out_dir: Path) -> list:
    """Search YouTube and download up to `per` short audio clips. Returns [wav paths]."""
    try:
        from yt_dlp import YoutubeDL
    except ImportError:
        print("[fatal] yt-dlp not installed", flush=True)
        return []
    query = f"ytsearch{per*2}:{name} قرآن تلاوة"
    outtmpl = str(out_dir / f"%(id)s.%(ext)s")
    ydl_opts = {
        "quiet": True,
        "no_warnings": True,
        "format": "bestaudio/best",
        "outtmpl": outtmpl,
        "noplaylist": True,
        "max_downloads": per * 2,
        "postprocessors": [{
            "key": "FFmpegExtractAudio",
            "preferredcodec": "wav",
            "preferredquality": "0",
        }],
        "postprocessor_args": [
            "-ac", "1", "-ar", "16000",
            "-ss", str(CLIP_OFFSET),
            "-t", str(CLIP_SECONDS),
        ],
        "socket_timeout": 30,
        "retries": 2,
    }
    results = []
    try:
        with YoutubeDL(ydl_opts) as ydl:
            ydl.download([query])
    except Exception as e:
        print(f"  [yt] {name}: {e}", flush=True)
    for f in out_dir.glob("*.wav"):
        results.append(f)
        if len(results) >= per:
            break
    return results


def main():
    print("=" * 70, flush=True)
    print("YouTube enrichment for missing reciters", flush=True)
    print("=" * 70, flush=True)

    names = _load_excel_names()
    if not names:
        print("[exit] no excel names", flush=True); return
    print(f"Excel: {len(names)} reciters", flush=True)

    existing = _load_existing_bases()
    print(f"DB already covers: {len(existing)} unique reciters", flush=True)

    missing = [n for n in names if _norm(n) not in existing]
    print(f"Missing from DB: {len(missing)} reciters", flush=True)
    if not missing:
        return

    # Load AI engine
    print("Loading AI engine (ensemble ECAPA + xvect)...", flush=True)
    from app.ai_engine import VoiceRecognitionEngine
    eng = VoiceRecognitionEngine()

    # Load current DB
    if DB_PATH.exists():
        with open(DB_PATH, encoding="utf-8") as f:
            db = json.load(f)
    else:
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        db = {"version": "3.0-ensemble", "reciters": []}

    if META_PATH.exists():
        with open(META_PATH, encoding="utf-8") as f:
            meta = json.load(f)
    else:
        meta = {"reciters": []}

    added_total = 0
    for i, name in enumerate(missing, 1):
        print(f"\n[{i}/{len(missing)}] {name}", flush=True)
        with tempfile.TemporaryDirectory() as td:
            wavs = _yt_search_and_download(name, PER_RECITER, Path(td))
            if not wavs:
                print("  [skip] no downloads", flush=True); continue
            added = 0
            for j, wav in enumerate(wavs):
                try:
                    if not eng.validate_audio_duration(wav, min_duration_sec=15):
                        continue
                    emb = eng.process_audio_file(wav)
                    key = f"{name}#{j}"
                    db["reciters"].append({
                        "reciter_name": key,
                        "embedding": emb.tolist(),
                        "source": "youtube",
                    })
                    added += 1
                except Exception as e:
                    print(f"  [err] sample {j}: {e}", flush=True)
            if added:
                meta["reciters"].append({
                    "name": name, "name_english": name, "country": "-",
                    "bio": f"مضاف من YouTube — {added} عيّنة",
                    "birth_year": "-", "death_year": None, "image_url": "",
                    "recitation_style": "مرتّل",
                })
                added_total += added
                print(f"  ✓ +{added} fingerprints", flush=True)
                # Save periodically
                if i % 5 == 0:
                    _save(db, meta)
        time.sleep(1)  # polite

    _save(db, meta)
    print(f"\n{'='*70}\nDONE — added {added_total} fingerprints", flush=True)
    print(f"DB now has {len(db['reciters'])} fingerprints", flush=True)


def _save(db, meta):
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    META_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False)
    with open(META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("interrupted")
