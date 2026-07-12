"""Helpers for youtube enrichment."""
import csv, json, os, re, unicodedata
from pathlib import Path
B = Path(__file__).resolve().parent.parent
DB = B/"data"/"embeddings"/"reciter_database.json"
MT = B/"data"/"reciters_metadata.json"
RJ = B/"data"/"youtube_enrichment_report.json"
RC = B/"data"/"youtube_enrichment_report.csv"
XP = Path(os.environ.get("QURAN_EXCEL", B/"data"/"reciters.xlsx"))

def norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"[\u0625\u0623\u0622\u0627]","\u0627",s)
    s = re.sub(r"[\u0649\u064a]","\u064a",s)
    s = re.sub(r"\u0629","\u0647",s)
    s = re.sub(r"[\u064b-\u0652\u0640]","",s)
    return re.sub(r"\s+"," ",s).strip().lower()

def bases():
    if not DB.exists(): return set()
    d = json.load(open(DB,encoding="utf-8"))
    return {norm(r["reciter_name"].split("#")[0]) for r in d.get("reciters",[])}

def excel_names():
    if not XP.exists(): return []
    from openpyxl import load_workbook
    ws = load_workbook(str(XP), data_only=True).active
    names, seen = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        for c in row:
            if c and isinstance(c,str) and len(c.strip())>2:
                k = norm(c)
                if k and k not in seen: seen.add(k); names.append(c.strip())
                break
    return names

def save(db, meta, rep):
    for p in (DB,MT,RJ,RC): p.parent.mkdir(parents=True, exist_ok=True)
    json.dump(db, open(DB,"w",encoding="utf-8"), ensure_ascii=False)
    json.dump(meta, open(MT,"w",encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(rep, open(RJ,"w",encoding="utf-8"), ensure_ascii=False, indent=2)
    with open(RC,"w",encoding="utf-8",newline="") as f:
        w = csv.writer(f)
        w.writerow(["reciter","status","fingerprints_added","videos_tried","video_urls","video_titles","error"])
        for r in rep["reciters"]:
            w.writerow([r["name"], r["status"], r["fingerprints_added"], len(r["videos"]),
                " | ".join(v["url"] for v in r["videos"]),
                " | ".join(v.get("title","") for v in r["videos"]),
                r.get("error","")])
